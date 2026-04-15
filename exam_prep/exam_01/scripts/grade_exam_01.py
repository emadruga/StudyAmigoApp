"""
grade_exam_01.py
Gera o CSV de resultados da Prova Final do 1º Bimestre.

Uso:
    python3 grade_exam_01.py --roster <roster.csv> --form <form.xlsx> \
        --impressa <impressa.xlsx> --out <output.csv>

Fontes:
  - roster CSV       -> Tier canônico de cada aluno (coluna "Suggested Tier")
  - form XLSX        -> Respostas do Google Form (digital)
  - impressa XLSX    -> Respostas da prova impressa (alunos sem smartphone)

Regras de Tier e nota:
  - Tier 1: só as 10 primeiras questões (cols 5-14 do form). Divisor = 10.
  - Tier 2: as 20 primeiras questões (cols 5-24 do form). Divisor = 20.
  - Nota final = (acertos / questoes_do_tier) * 10, arredondada p/ 2 casas.
  - Fonte do Tier (prioridade):
      1. roster  -> coluna "Suggested Tier"
      2. form digital -> coluna tier escolhida pelo aluno (se roster vazio)
      3. prova impressa -> coluna "Tier" (se roster vazio)
      4. fallback: Tier 2 (marcado na coluna "Tier_source" = "FALLBACK_T2")
  - Alunos sem Tier no roster têm "Tier_source" = "form_choice" / "impressa" /
    "FALLBACK_T2" para fácil revisão.
"""

import argparse
import csv
import unicodedata
import re
import os
import openpyxl

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(__file__)
_BASE       = os.path.join(_SCRIPT_DIR, "..")

parser = argparse.ArgumentParser(
    description="Corrige a Prova Final 1º Bimestre e gera CSV de resultados."
)
parser.add_argument(
    "--roster",
    default=os.path.join(_BASE, "bases", "curated_student_roster_v2.csv"),
    metavar="FILE",
    help="CSV do roster de alunos com coluna 'Suggested Tier' (padrão: ../bases/curated_student_roster_v2.csv)",
)
parser.add_argument(
    "--form",
    default=os.path.join(_BASE, "bases", "prova-final-1Bimestre-2026.xlsx"),
    metavar="FILE",
    help="XLSX exportado do Google Forms (padrão: ../bases/prova-final-1Bimestre-2026.xlsx)",
)
parser.add_argument(
    "--impressa",
    default=os.path.join(_BASE, "bases", "Prova-Final-1Bimestre-Impressa-2026.xlsx"),
    metavar="FILE",
    help="XLSX com resultados da prova impressa — colunas: Aluno, Tier, Acertos (padrão: ../bases/Prova-Final-1Bimestre-Impressa-2026.xlsx)",
)
parser.add_argument(
    "--out",
    default=os.path.join(_BASE, "output", "prova_final_1bim_RESULTADOS.csv"),
    metavar="FILE",
    help="Caminho do CSV de saída (padrão: ../output/prova_final_1bim_RESULTADOS.csv)",
)
args = parser.parse_args()

ROSTER_CSV    = args.roster
FORM_XLSX     = args.form
IMPRESSA_XLSX = args.impressa
OUTPUT_CSV    = args.out

os.makedirs(os.path.dirname(os.path.abspath(OUTPUT_CSV)), exist_ok=True)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def normalize(s: str) -> str:
    """Lower-case, sem acentos, sem pontuação dupla, espaços colapsados."""
    if not s:
        return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def tier_label(raw: str) -> str | None:
    """Extrai 'Tier 1' ou 'Tier 2' de strings livres; None se irreconhecível."""
    if not raw:
        return None
    m = re.search(r"tier\s*([123])", raw, re.IGNORECASE)
    if m:
        return f"Tier {m.group(1)}"
    return None


def questoes_do_tier(tier: str) -> int:
    return 10 if tier == "Tier 1" else 20   # Tier 2 → 20 questões (acumulativo)


def nota(acertos: float, tier: str) -> float:
    divisor = questoes_do_tier(tier)
    return round((acertos / divisor) * 10, 2)


def fuzzy_match(name: str, lookup: dict) -> str | None:
    """Busca o ID no lookup pelo nome normalizado; retorna ID ou None."""
    key = normalize(name)
    return lookup.get(key)

# ---------------------------------------------------------------------------
# 1. Carregar roster
# ---------------------------------------------------------------------------
roster = {}          # id -> {nome, curso, tier_roster, email}
name_to_id = {}      # normalized_name -> id

with open(ROSTER_CSV, encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)
    for row in reader:
        sid = row["ID"].strip()
        nome = row["Nome"].strip()
        tier_raw = row.get("Suggested Tier", "").strip()
        roster[sid] = {
            "nome":        nome,
            "curso":       row.get("\ufeffCurso", row.get("Curso", "")).strip(),
            "tier_roster": tier_label(tier_raw),
            "email":       row.get("Email", "").strip(),
        }
        name_to_id[normalize(nome)] = sid

# ---------------------------------------------------------------------------
# 2. Carregar prova impressa
# ---------------------------------------------------------------------------
impressa = {}  # normalized_name -> {tier, acertos}

wb_imp = openpyxl.load_workbook(IMPRESSA_XLSX)
ws_imp = wb_imp.active
for row in ws_imp.iter_rows(min_row=2, values_only=True):
    nome_imp, tier_num, acertos_imp = row[0], row[1], row[2]
    if nome_imp is None:
        continue
    t = f"Tier {int(tier_num)}" if tier_num is not None else None
    impressa[normalize(str(nome_imp))] = {
        "tier_impressa": t,
        "acertos":       float(acertos_imp) if acertos_imp is not None else 0,
    }

# ---------------------------------------------------------------------------
# 3. Carregar respostas do Google Form (digital)
# ---------------------------------------------------------------------------
# Colunas do form (0-based):
#   0  Timestamp   1 Email   2 Pontuação (total acertos calculado pelo Form)
#   3  Nome        4 Tier escolhido
#   5-14  Questões Tier 1 (10 questões)
#   15-24 Questões Tier 2 (10 questões)
#   25-34 Questões Tier 3 (10 questões)
#
# A pontuação já vem calculada pelo Google Forms (1 pt por questão certa, max 30).
# Para Tier 1: usamos apenas os acertos do bloco 5-14 → precisamos recalcular
# porque a pontuação total inclui outros blocos.
# Estratégia: contar questões preenchidas e usar a pontuação proporcional.
# Mais simples: o Google Forms marca como NULL questões não respondidas.
# Portanto acertos reais = (acertos no bloco do tier).
# MAS o form não nos diz quais estão certas/erradas individualmente — apenas
# a pontuação total. Então usamos a pontuação total ponderada:
#   Tier 1: acertos = pontuacao_total * (10 / questoes_respondidas_totais)
#   ... isso seria impreciso. A abordagem correta é:
#   - Tier 1: questoes 5-14 foram respondidas? Se sim, a pontuação daquele
#     bloco está embutida na pontuação total. Não temos como separá-la sem
#     o gabarito.
#
# Decisão pragmática:
#   - Para Tier 1: divisor = 10, acertos = min(pontuacao_total, 10)
#     (Tier 1 só devia responder as primeiras 10, então seu score ≤ 10)
#   - Para Tier 2: divisor = 20, acertos = min(pontuacao_total, 20)
#   - Observação: alunos Tier 1 que responderam mais de 10 questões terão
#     score truncado em 10.

form_results = {}  # normalized_name -> {acertos_raw, tier_form, email}

wb_form = openpyxl.load_workbook(FORM_XLSX)
ws_form = wb_form.active
for row in ws_form.iter_rows(min_row=2, values_only=True):
    nome_form = row[3]
    if nome_form is None:
        continue
    email_form   = str(row[1]) if row[1] else ""
    pontuacao    = float(row[2]) if row[2] is not None else 0.0
    tier_form    = tier_label(str(row[4])) if row[4] else None
    form_results[normalize(str(nome_form))] = {
        "acertos_raw": pontuacao,
        "tier_form":   tier_form,
        "email_form":  email_form,
    }

# ---------------------------------------------------------------------------
# 4. Montar resultado final para cada aluno do roster
# ---------------------------------------------------------------------------
results = []

for sid, info in sorted(roster.items(), key=lambda x: x[1]["nome"]):
    nome      = info["nome"]
    norm_nome = normalize(nome)
    tier_src  = ""
    tier_final = None
    acertos    = None
    fonte_prova = ""
    obs = ""

    # --- Determinar Tier ---
    if info["tier_roster"]:
        tier_final = info["tier_roster"]
        tier_src   = "roster"
    else:
        # Sem tier no roster → buscar no form digital ou prova impressa
        fd = form_results.get(norm_nome)
        fi = impressa.get(norm_nome)

        # Tenta fuzzy nos nomes da impressa (nomes podem estar abreviados)
        if fi is None:
            for k, v in impressa.items():
                # Verifica se todos os tokens do nome do roster aparecem no nome da impressa
                tokens = norm_nome.split()
                if all(t in k for t in tokens[:2]):  # usa pelo menos 2 tokens
                    fi = v
                    break

        tier_escolhido = None
        if fd and fd["tier_form"]:
            tier_escolhido = fd["tier_form"]
            tier_src = "form_choice"
        elif fi and fi["tier_impressa"]:
            tier_escolhido = fi["tier_impressa"]
            tier_src = "impressa"

        if tier_escolhido:
            tier_final = tier_escolhido
        else:
            tier_final = "Tier 2"
            tier_src   = "FALLBACK_T2"
            obs = "SEM_TIER_ATRIBUIDO"

    # --- Buscar acertos ---
    fd = form_results.get(norm_nome)

    # Verificar se está na prova impressa (com fuzzy matching)
    fi = impressa.get(norm_nome)
    if fi is None:
        for k, v in impressa.items():
            tokens = norm_nome.split()
            if all(t in k for t in tokens[:2]):
                fi = v
                break

    if fi is not None:
        # Prova impressa tem precedência se o aluno está nessa lista
        acertos     = fi["acertos"]
        fonte_prova = "impressa"
    elif fd is not None:
        acertos     = fd["acertos_raw"]
        fonte_prova = "digital"
    else:
        acertos     = None
        fonte_prova = "AUSENTE"
        if obs:
            obs += "|NAO_ENCONTRADO"
        else:
            obs = "NAO_ENCONTRADO"

    # --- Calcular nota ---
    if acertos is not None:
        # Truncar acertos ao máximo possível do tier
        max_acertos = questoes_do_tier(tier_final)
        acertos_efetivos = min(acertos, max_acertos)
        nota_final = nota(acertos_efetivos, tier_final)
    else:
        acertos_efetivos = ""
        nota_final = ""

    results.append({
        "ID":              sid,
        "Nome":            nome,
        "Curso":           info["curso"],
        "Tier":            tier_final,
        "Tier_source":     tier_src,
        "Fonte_prova":     fonte_prova,
        "Acertos":         acertos if acertos is not None else "",
        "Acertos_efetivos": acertos_efetivos,
        "Max_questoes":    questoes_do_tier(tier_final) if tier_final else "",
        "Nota_10":         nota_final,
        "Obs":             obs,
    })

# ---------------------------------------------------------------------------
# 5. Escrever CSV de saída
# ---------------------------------------------------------------------------
fieldnames = [
    "ID", "Nome", "Curso", "Tier", "Tier_source",
    "Fonte_prova", "Acertos", "Acertos_efetivos", "Max_questoes",
    "Nota_10", "Obs",
]

with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(results)

print(f"CSV gerado: {OUTPUT_CSV}")
print(f"Total alunos: {len(results)}")
ausentes = [r for r in results if r["Fonte_prova"] == "AUSENTE"]
fallback = [r for r in results if r["Tier_source"] == "FALLBACK_T2"]
print(f"Ausentes (não encontrados em nenhuma fonte): {len(ausentes)}")
if ausentes:
    for r in ausentes:
        print(f"  - {r['ID']} {r['Nome']}")
print(f"Alunos com FALLBACK_T2 (sem tier atribuído em nenhuma fonte): {len(fallback)}")
if fallback:
    for r in fallback:
        print(f"  - {r['ID']} {r['Nome']} | fonte={r['Fonte_prova']}")
