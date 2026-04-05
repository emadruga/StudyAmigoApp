#!/usr/bin/env python3
"""
add_post_e01_sheet.py

Reads curated_student_roster.xlsx (strict OOXML) and adds a new sheet
"curated_student_roster_apos_E01" with:
  - All existing roster students (from sheet1)
  - Extra students found in E01_final_grades.csv who have no student_id
    (i.e. did not take the placement exam)
  - nota_E01 column (0-10) filled from E01_final_grades.csv where available

New sheet columns:
  Curso | ID | Nome | Email | Caminho | Suggested Tier | nota_E01 | obs_E01

Usage:
    python placement_exam/planning_E01/scripts/add_post_e01_sheet.py
"""

import csv
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

XLSX_IN    = Path("/Users/emadruga/Desktop/Aulas CtSegCiber/2026-1/Lingua Inglesa/curated_student_roster.xlsx")
XLSX_OUT   = Path("/Users/emadruga/Desktop/Aulas CtSegCiber/2026-1/Lingua Inglesa/curated_student_roster.xlsx")
GRADES_CSV = Path("placement_exam/planning_E01/E01_final_grades.csv")
ROSTER_CSV = Path("placement_exam/bases/curated_student_roster.csv")  # used to seed seen_names
NEW_SHEET  = "curated_student_roster_apos_E01"

# Columns in existing sheet (Portuguese headers as found in xlsx)
EXISTING_HEADERS = ["Curso", "ID", "Nome", "Email", "Caminho", "Suggested Tier"]
NEW_HEADERS = EXISTING_HEADERS + ["nota_E01", "obs_E01"]

# ---------------------------------------------------------------------------
# 1. Read existing sheet — tries openpyxl first, falls back to raw XML
#    (needed because the original file used strict OOXML namespace)
# ---------------------------------------------------------------------------
def read_existing_sheet(xlsx_path):
    """Returns list of dicts with keys matching EXISTING_HEADERS."""
    # Try openpyxl (works after first save)
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.worksheets[0]
        result = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            vals = [str(v) if v is not None else '' for v in row]
            while len(vals) < len(EXISTING_HEADERS):
                vals.append('')
            result.append(dict(zip(EXISTING_HEADERS, vals)))
        if result:
            return result
    except Exception:
        pass

    # Fallback: raw XML for strict OOXML files
    ns = '{http://purl.oclc.org/ooxml/spreadsheetml/main}'
    with zipfile.ZipFile(xlsx_path) as z:
        names = z.namelist()
        if 'xl/sharedStrings.xml' in names:
            ss_root = ET.fromstring(z.read('xl/sharedStrings.xml').decode())
            strings = [
                si.findtext(f'{ns}t') or ''.join(t.text or '' for t in si.iter(f'{ns}t'))
                for si in ss_root
            ]
        else:
            strings = []
        s1_root = ET.fromstring(z.read('xl/worksheets/sheet1.xml').decode())

    rows = []
    for row in s1_root.iter(f'{ns}row'):
        r = []
        for c in row:
            t = c.get('t', '')
            v_el = c.find(f'{ns}v')
            val = v_el.text if v_el is not None else ''
            if t == 's' and val and strings:
                val = strings[int(val)]
            r.append(val)
        rows.append(r)

    result = []
    for row in rows[1:]:
        if not any(row):
            continue
        while len(row) < len(EXISTING_HEADERS):
            row.append('')
        result.append(dict(zip(EXISTING_HEADERS, row)))
    return result

# ---------------------------------------------------------------------------
# 2. Read E01 final grades
# ---------------------------------------------------------------------------
def read_grades(csv_path):
    """Returns dict keyed by student_id (str) and by name (lower) → grade row."""
    by_id   = {}
    by_name = {}
    with open(csv_path, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            sid  = (row.get('student_id') or '').strip()
            name = (row.get('name') or '').strip()
            nota = (row.get('nota_final') or '0.0').strip()
            flags = (row.get('flags') or '').strip()
            uid  = (row.get('user_id') or '').strip()
            username = (row.get('username') or '').strip()
            email = (row.get('email') or '').strip()
            course = (row.get('course') or '').strip()
            tier = (row.get('tier') or '').strip()
            path = (row.get('path') or '').strip()

            entry = {
                'nota_E01': nota,
                'flags':    flags,
                'user_id':  uid,
                'username': username,
                'email':    email,
                'course':   course,
                'tier':     tier,
                'path':     path,
                'name':     name,
            }
            if sid:
                by_id[sid] = entry
            if name:
                by_name[name.lower()] = entry
    return by_id, by_name

# ---------------------------------------------------------------------------
# 3. Normalize course name
# ---------------------------------------------------------------------------
COURSE_MAP = {
    'segurança cibernética': 'SegCiber',
    'segcyber': 'SegCiber',
    'segciber': 'SegCiber',
    'metrologia': 'Metrologia',
    'biotecnologia': 'Biotecnologia',
}
def norm_course(c):
    return COURSE_MAP.get(c.lower().strip(), c)

# ---------------------------------------------------------------------------
# 4. Build obs string from flags
# ---------------------------------------------------------------------------
FLAG_LABELS = {
    'RET100':    'Retenção 100% (verificar)',
    'LOW_TIME':  'Tempo baixo por resposta',
    'CRAM':      'Cramming no último dia',
    'NO_ACCOUNT':'Sem conta no app',
}
def flags_to_obs(flags_str):
    parts = [FLAG_LABELS.get(f.strip(), f.strip()) for f in flags_str.split(',') if f.strip()]
    return '; '.join(parts)

# ---------------------------------------------------------------------------
# 5. Main
# ---------------------------------------------------------------------------
def main():
    existing = read_existing_sheet(XLSX_IN)
    by_id, by_name = read_grades(GRADES_CSV)

    # Build rows for existing roster students
    output_rows = []
    # Seed seen_names from the CSV roster (reliable, correct encoding)
    seen_names = set()
    with open(ROSTER_CSV, encoding='latin-1', newline='') as _f:
        for _row in csv.DictReader(_f):
            _n = (_row.get('Name') or '').strip()
            if _n:
                seen_names.add(_n.lower())

    for s in existing:
        sid  = s['ID'].strip()
        name = s['Nome'].strip()
        seen_names.add(name.lower())

        grade_info = by_id.get(sid) or by_name.get(name.lower()) or {}
        nota  = grade_info.get('nota_E01', '')
        flags = grade_info.get('flags', '')
        obs   = flags_to_obs(flags)

        # If NO_ACCOUNT and nota is '' or 0.0, make it explicit
        if 'NO_ACCOUNT' in flags:
            nota = '0.0'

        output_rows.append({
            'Curso':          s['Curso'],
            'ID':             sid,
            'Nome':           name,
            'Email':          s['Email'],
            'Caminho':        s['Caminho'],
            'Suggested Tier': s['Suggested Tier'],
            'nota_E01':       nota,
            'obs_E01':        obs,
        })

    # Add extra students from grades CSV who have no student_id
    # (did not take placement exam — not in roster)
    SKIP_NAMES = {'test user', 'leonardo madruga', 'rayssa assis'}  # system/test accounts

    with open(GRADES_CSV, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            sid  = (row.get('student_id') or '').strip()
            name = (row.get('name') or '').strip()
            if sid:
                continue  # already in roster block above
            if not name or name.lower() in SKIP_NAMES:
                continue
            if name.lower() in seen_names:
                continue
            nota  = (row.get('nota_final') or '0.0').strip()
            flags = (row.get('flags') or '').strip()
            obs   = flags_to_obs(flags) + (' | Sem nivelamento' if not flags else '; Sem nivelamento')
            email  = (row.get('email') or '').strip()
            course = norm_course(row.get('course') or '')
            tier   = (row.get('tier') or '').strip()
            path   = (row.get('path') or '').strip()

            seen_names.add(name.lower())
            output_rows.append({
                'Curso':          course,
                'ID':             '',   # no ID yet
                'Nome':           name,
                'Email':          email,
                'Caminho':        path,
                'Suggested Tier': tier,
                'nota_E01':       nota,
                'obs_E01':        obs,
            })

    # ---------------------------------------------------------------------------
    # Write new sheet into the xlsx
    # ---------------------------------------------------------------------------
    wb = openpyxl.Workbook()
    # Recreate original sheet (existing already read above)
    ws1 = wb.active
    ws1.title = "curated_student_roster"
    ws1.append(EXISTING_HEADERS)
    for s in existing:
        ws1.append([s.get(h, '') for h in EXISTING_HEADERS])

    # Style header row of sheet1
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Create new sheet
    ws2 = wb.create_sheet(title=NEW_SHEET)
    ws2.append(NEW_HEADERS)

    # Style header
    header_fill2 = PatternFill("solid", fgColor="375623")
    for cell in ws2[1]:
        cell.fill = header_fill2
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')

    # Row fill colours
    fill_no_activity = PatternFill("solid", fgColor="FFE0E0")   # light red — zero/no account
    fill_flag        = PatternFill("solid", fgColor="FFF2CC")   # light yellow — flagged
    fill_no_id       = PatternFill("solid", fgColor="DDEBF7")   # light blue — no placement exam

    for r in output_rows:
        nota_str = r.get('nota_E01', '')
        obs_str  = r.get('obs_E01', '')
        sid      = r.get('ID', '')

        try:
            nota_val = float(nota_str) if nota_str else 0.0
        except ValueError:
            nota_val = 0.0

        row_data = [
            r['Curso'],
            int(sid) if sid.isdigit() else sid,
            r['Nome'],
            r['Email'],
            r['Caminho'],
            r['Suggested Tier'],
            nota_val,
            obs_str,
        ]
        ws2.append(row_data)

        # Colour coding
        last_row = ws2.max_row
        if not sid:
            row_fill = fill_no_id
        elif nota_val == 0.0:
            row_fill = fill_no_activity
        elif obs_str and any(f in obs_str for f in ['Retenção', 'Tempo baixo', 'Cramming']):
            row_fill = fill_flag
        else:
            row_fill = None

        if row_fill:
            for col in range(1, len(NEW_HEADERS) + 1):
                ws2.cell(row=last_row, column=col).fill = row_fill

    # Column widths
    col_widths = [14, 8, 38, 38, 10, 36, 10, 40]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # nota_E01 column: number format
    nota_col = NEW_HEADERS.index('nota_E01') + 1
    for row in ws2.iter_rows(min_row=2, min_col=nota_col, max_col=nota_col):
        for cell in row:
            cell.number_format = '0.0'

    wb.save(XLSX_OUT)
    print(f"Saved: {XLSX_OUT}")
    print(f"Nova aba '{NEW_SHEET}': {len(output_rows)} alunos")
    print(f"  - Com ID (do roster):      {sum(1 for r in output_rows if r['ID'])}")
    print(f"  - Sem ID (sem nivelamento):{sum(1 for r in output_rows if not r['ID'])}")
    print(f"  - Nota > 0:                {sum(1 for r in output_rows if float(r.get('nota_E01') or 0) > 0)}")
    print(f"  - Nota = 0:                {sum(1 for r in output_rows if float(r.get('nota_E01') or 0) == 0)}")

if __name__ == "__main__":
    main()
