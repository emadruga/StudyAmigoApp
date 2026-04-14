#!/usr/bin/env python3
"""
build_roster_xlsx.py — Gera curated_student_roster_E02.xlsx

Combina:
  - XLSX de entrada com notas E01 (--input)
  - CSV de notas E02 gerado por grade_exercise_v2 (--grades)
  - CSV do roster v2 com alunos novos a inserir (--roster-v2, opcional)

Saída:
  - XLSX com as mesmas colunas do xlsx de entrada + E02 + obs_E02 (--output)
  - Linhas ordenadas alfabeticamente por Nome

Matching:
  O CSV usa student_id (matrícula) como chave primária. Quando student_id
  está ausente ou vazio no CSV, o script tenta fuzzy-match pelo nome.

Uso:
  python placement_exam/planning_E02/scripts/build_roster_xlsx.py \\
      --input     placement_exam/planning_E02/bases/curated_student_roster.xlsx \\
      --grades    placement_exam/planning_E02/output/E02_final_grades.csv \\
      --roster-v2 exam_prep/exam_01/bases/curated_student_roster_v2.csv \\
      --output    placement_exam/planning_E02/output/curated_student_roster_E02.xlsx
"""

import argparse
import csv
import difflib
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FUZZY_THRESHOLD = 0.70  # mínimo para aceitar match por nome


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _norm(name: str) -> str:
    """Lower-case, sem acentos básicos, sem pontuação extra."""
    name = name.lower().strip()
    for src, dst in [
        ("á","a"),("à","a"),("ã","a"),("â","a"),
        ("é","e"),("ê","e"),
        ("í","i"),
        ("ó","o"),("ô","o"),("õ","o"),
        ("ú","u"),("ü","u"),
        ("ç","c"),
    ]:
        name = name.replace(src, dst)
    return re.sub(r"[^a-z0-9 ]", "", name)


def load_grades(grades_path: Path):
    """
    Retorna dois dicionários:
        by_id   : {student_id_str -> row_dict}
        by_name : {normalised_name -> row_dict}   (apenas linhas com grade > 0
                                                    ou zero-review estudantes
                                                    no roster)
    """
    by_id: dict = {}
    by_name: dict = {}
    with open(grades_path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            sid = row.get("student_id", "").strip()
            name = row.get("name", "").strip()
            if sid and sid != "NO_ACCOUNT":
                by_id[sid] = row
            if name:
                key = _norm(name)
                # keep highest-grade entry when there are multiple accounts
                if key not in by_name or float(row["grade"]) > float(by_name[key]["grade"]):
                    by_name[key] = row
    return by_id, by_name


def find_grade_row(student_id, student_name: str, by_id: dict, by_name: dict):
    """Tenta encontrar a linha de nota pelo ID; cai para fuzzy por nome."""
    sid = str(student_id) if student_id is not None else ""
    if sid and sid in by_id:
        return by_id[sid], "id"

    # fuzzy por nome
    key = _norm(student_name)
    best_score = 0.0
    best_row = None
    for cand_key, row in by_name.items():
        score = difflib.SequenceMatcher(None, key, cand_key).ratio()
        if score > best_score:
            best_score = score
            best_row = row
    if best_score >= FUZZY_THRESHOLD:
        return best_row, f"fuzzy:{best_score:.2f}"
    return None, None


def grade_letter(grade_val):
    """Converte nota numérica para letra (mesmo critério do grade_exercise_v2)."""
    g = float(grade_val)
    if g >= 90: return "A+"
    if g >= 80: return "A"
    if g >= 70: return "B"
    if g >= 60: return "C"
    if g >= 50: return "D"
    return "F"


def build_obs(row: dict) -> str:
    """Monta string de observação para obs_E02 a partir dos flags e métricas."""
    parts = []
    flags = row.get("flags", "").strip()
    if flags:
        parts.append(flags)
    if row.get("username", "").strip() and row.get("username") != row.get("name"):
        parts.append(f'username: {row["username"]}')
    return " | ".join(parts) if parts else None


# ─────────────────────────────────────────────────────────────────────────────
# Styling helpers
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # azul escuro
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)

CURSO_COLORS = {
    "biotecnologia": "D9E1F2",   # azul-cinza claro
    "metrologia":    "E2EFDA",   # verde claro
    "segciber":      "FCE4D6",   # laranja claro
}

ZERO_FILL    = PatternFill("solid", fgColor="FFD7D7")   # rosa para nota 0
LOW_FILL     = PatternFill("solid", fgColor="FFF2CC")   # amarelo para nota < 5
FLAG_FONT    = Font(color="C00000", bold=True)           # vermelho para flags


def _curso_fill(curso: str):
    key = _norm(curso).replace(" ", "")
    # match parcial
    for k, color in CURSO_COLORS.items():
        if k in key:
            return PatternFill("solid", fgColor=color)
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input",      required=True, metavar="XLSX",
                        help="Planilha de entrada com notas E01 (ex.: bases/curated_student_roster.xlsx)")
    parser.add_argument("--grades",     required=True, metavar="CSV",
                        help="CSV de notas E02 gerado por grade_exercise_v2 (ex.: output/E02_final_grades.csv)")
    parser.add_argument("--output",     required=True, metavar="XLSX",
                        help="Planilha de saída (ex.: output/curated_student_roster_E02.xlsx)")
    parser.add_argument("--roster-v2",  default=None,  metavar="CSV",
                        help="CSV do roster v2 para inserir alunos ausentes do xlsx de entrada (opcional)")
    args = parser.parse_args()

    input_path    = Path(args.input)
    grades_path   = Path(args.grades)
    output_path   = Path(args.output)
    roster_v2_path = Path(args.roster_v2) if args.roster_v2 else None

    # Validações
    for p in (input_path, grades_path):
        if not p.exists():
            sys.exit(f"Arquivo não encontrado: {p}")
    if roster_v2_path and not roster_v2_path.exists():
        sys.exit(f"Arquivo não encontrado (--roster-v2): {roster_v2_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ── Carregar notas E02 ────────────────────────────────────────────────────
    print(f"Carregando notas E02: {grades_path}")
    by_id, by_name = load_grades(grades_path)
    print(f"  {len(by_id)} entradas com student_id, {len(by_name)} entradas por nome")

    # ── Carregar xlsx de entrada ──────────────────────────────────────────────
    print(f"Carregando roster: {input_path}")
    wb_in = openpyxl.load_workbook(input_path)

    # ── Criar workbook de saída ───────────────────────────────────────────────
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # remove aba padrão

    # A aba "E01" contém as colunas Tier + E01 + obs_E01 que queremos preservar.
    # Usamos ela como fonte principal para a aba de saída.
    SOURCE_SHEET = "E01" if "E01" in wb_in.sheetnames else wb_in.sheetnames[0]
    print(f"  Aba fonte para E01: '{SOURCE_SHEET}'")

    matched = 0
    unmatched_names = []

    for sheet_name in wb_in.sheetnames:
        ws_in  = wb_in[SOURCE_SHEET if sheet_name == wb_in.sheetnames[0] else sheet_name]
        out_title = sheet_name  # mantém nome original da aba na saída
        ws_out = wb_out.create_sheet(title=out_title)

        # Detectar linha de cabeçalho (primeira linha com "ID" ou "Nome")
        header_row_idx = None
        for i, row in enumerate(ws_in.iter_rows(values_only=True), start=1):
            if row and any(str(c or "").strip().lower() in ("id", "nome", "name") for c in row):
                header_row_idx = i
                break

        if header_row_idx is None:
            # Aba sem cabeçalho reconhecível — copiar como está
            for row in ws_in.iter_rows():
                for cell in row:
                    ws_out[cell.coordinate].value = cell.value
            continue

        # Ler cabeçalho original
        orig_headers = [str(c or "").strip() for c in
                        next(ws_in.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))]

        # Novo cabeçalho: adiciona E02 e obs_E02
        new_headers = orig_headers + ["E02", "obs_E02"]

        # Mapear índices das colunas-chave (0-based)
        col_idx = {h.lower().rstrip(): i for i, h in enumerate(orig_headers)}
        id_col   = col_idx.get("id")
        nome_col = col_idx.get("nome") or col_idx.get("name")

        # ── Escrever cabeçalho ────────────────────────────────────────────────
        for col_num, header in enumerate(new_headers, start=1):
            cell = ws_out.cell(row=1, column=col_num, value=header)
            cell.font  = HEADER_FONT
            cell.fill  = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Coletar linhas de dados em memória ───────────────────────────────────
        # Cada entrada: (nome_para_sort, new_row_vals, e02_grade, obs_e02)
        data_rows: list = []

        for row_vals in ws_in.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not any(v is not None for v in row_vals):
                continue  # pula linhas completamente vazias

            student_id   = row_vals[id_col]   if id_col   is not None else None
            student_name = row_vals[nome_col]  if nome_col is not None else ""
            student_name = str(student_name or "").strip()

            grade_row, match_type = find_grade_row(student_id, student_name, by_id, by_name)

            if grade_row is not None:
                e02_grade = round(float(grade_row["grade"]) / 10, 1)
                sid_in_csv = grade_row.get("student_id", "").strip()
                if sid_in_csv == "NO_ACCOUNT":
                    obs_e02 = "Sem conta no app"
                else:
                    obs_e02 = build_obs(grade_row)
                if sheet_name == wb_in.sheetnames[0]:
                    matched += 1
            else:
                e02_grade = 0.0
                obs_e02   = "Sem atividade no E02"
                if sheet_name == wb_in.sheetnames[0] and student_name:
                    unmatched_names.append(student_name)

            data_rows.append((student_name, list(row_vals) + [e02_grade, obs_e02], e02_grade, obs_e02))

        # ── Inserir alunos do roster v2 ausentes do xlsx de entrada ──────────────
        if sheet_name == wb_in.sheetnames[0]:
            if roster_v2_path and roster_v2_path.exists():
                written_ids = {str(row_vals[id_col]).strip()
                               for _, row_vals, _, _ in data_rows
                               if id_col is not None and row_vals[id_col] is not None}

                with open(roster_v2_path, newline="", encoding="utf-8-sig") as fv2:
                    for v2row in csv.DictReader(fv2):
                        norm = {k.strip().lower(): v.strip() for k, v in v2row.items()}
                        sid = norm.get("id", "")
                        if not sid or sid in written_ids:
                            continue
                        name_v2  = norm.get("nome", norm.get("name", ""))
                        curso_v2 = norm.get("curso", norm.get("course", ""))
                        email_v2 = norm.get("email", "")
                        path_v2  = norm.get("caminho", norm.get("path", ""))
                        tier_v2  = norm.get("suggested tier", norm.get("tier", ""))
                        row_v2 = []
                        for h in orig_headers:
                            hl = h.lower().rstrip()
                            if hl == "curso":                  row_v2.append(curso_v2)
                            elif hl == "id":                   row_v2.append(int(sid) if sid.isdigit() else sid)
                            elif hl == "nome":                 row_v2.append(name_v2)
                            elif hl == "email":                row_v2.append(email_v2)
                            elif hl == "caminho":              row_v2.append(path_v2)
                            elif hl in ("tier", "tier "):      row_v2.append(tier_v2)
                            elif hl == "e01":                  row_v2.append(0)
                            elif hl == "obs_e01":              row_v2.append("Sem dados E01")
                            else:                              row_v2.append(None)

                        grade_row_v2, _ = find_grade_row(sid, name_v2, by_id, by_name)
                        if grade_row_v2 is not None:
                            e02_v2 = round(float(grade_row_v2["grade"]) / 10, 1)
                            obs_v2 = build_obs(grade_row_v2)
                        else:
                            e02_v2 = 0.0
                            obs_v2 = "Sem atividade no E02"

                        data_rows.append((name_v2, row_v2 + [e02_v2, obs_v2], e02_v2, obs_v2))
                        matched += 1
                        print(f"  [roster_v2] Adicionado: {sid} {name_v2}")

        # ── Ordenar por nome (apenas aba principal) ───────────────────────────────
        is_main = (sheet_name == wb_in.sheetnames[0])
        if is_main:
            data_rows.sort(key=lambda t: _norm(t[0]))

        # ── Escrever linhas ordenadas ─────────────────────────────────────────────
        out_row = 2
        for _, new_row_vals, e02_grade, obs_e02 in data_rows:
            curso = str(new_row_vals[0] or "").strip() if new_row_vals else ""
            fill  = _curso_fill(curso)

            for col_num, val in enumerate(new_row_vals, start=1):
                cell = ws_out.cell(row=out_row, column=col_num, value=val)
                cell.alignment = Alignment(vertical="center")

                if fill:
                    cell.fill = fill

                e02_col = len(orig_headers) + 1
                if col_num == e02_col:
                    if e02_grade == 0.0:
                        cell.fill = ZERO_FILL
                    elif e02_grade < 5.0:
                        cell.fill = LOW_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = "0.0"

                obs_col = len(orig_headers) + 2
                if col_num == obs_col and obs_e02 and any(
                    flag in str(obs_e02) for flag in ("RET100", "LOW_TIME", "CRAM")
                ):
                    cell.font = FLAG_FONT

            out_row += 1

        # Ajustar largura das colunas
        for col_cells in ws_out.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws_out.column_dimensions[col_letter].width = min(max_len + 3, 45)

        # Congelar primeira linha
        ws_out.freeze_panes = "A2"

    # ── Salvar ────────────────────────────────────────────────────────────────
    wb_out.save(output_path)
    print(f"\nResultado salvo em: {output_path}")
    print(f"  Alunos matched:   {matched}")
    if unmatched_names:
        print(f"  Sem nota E02 ({len(unmatched_names)}): {', '.join(unmatched_names[:10])}"
              + (" ..." if len(unmatched_names) > 10 else ""))


if __name__ == "__main__":
    main()
